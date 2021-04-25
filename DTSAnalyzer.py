# -*- coding: UTF-8 -*-
#
# DTS问题单分析统计工具
# Bingo
# FAQ:
#   1. 问题单当前处理人为2个或以上不同组的组员时，会出现统计的问题单重复，故问题单流程中当前处理人不要走给两个人同时修改。
#

import os
import sys
import time
import json
import traceback
import pandas as pd
from decimal import Decimal
from collections import OrderedDict
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

__author__ = 'lwx382598'
__version__ = '3.1'

g_in_excel = 'DTS-IN.xlsx'
g_out_excel = 'DTS-OUT.xlsx'
g_di_excel = 'DI-DAILY.xlsx'
g_member_cnf = 'member.json'
g_version_cnf = 'version.json'
g_settings_cnf = 'settings.json'
# 团队所有成员
g_member_dict = {}
# 所有支持的版本
g_include_versions = []
# 过滤排除的版本
g_exclude_versions = []
# 只统计该版本号中组员外的其他人员的问题单
g_single_version = None
# DTS url根路径
g_root_url = ''
g_di_level = ['致命', '严重', '一般', '提示']
g_level_map = {'致命': '10', '严重': '3', '一般': '1', '提示': '0.1'}
# 归档过程
g_filing_list = ['CMO归档']
# 测试回归过程
g_regress_list = ['测试经理组织测试', '测试人员回归测试', '确认问题单']
# debug开关
g_debug_switch = False
# 替换中文名
g_change_name = True
# 今天的日期
g_today = time.strftime("%Y%m%d", time.localtime(int(time.time())))


def exit_delay(t=5):
    Logger.info("即将在%s秒后退出程序..." % t)
    time.sleep(t)
    sys.exit(0)


def about_info():
    about = """
    \tDTS分析工具 v%s
    \t\t——%s
    """ % (__version__, __author__)
    print(about)


class Logger:
    @classmethod
    def get_time(cls):
        ct = time.time()
        return '%s.%03d' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), (ct - int(ct)) * 1000)

    @classmethod
    def info(cls, info):
        print("[INFO ] %s: %s" % (cls.get_time(), str(info)))

    @classmethod
    def debug(cls, info):
        if not g_debug_switch:
            return
        print("[DEBUG] %s: %s" % (cls.get_time(), str(info)))

    @classmethod
    def error(cls, info):
        print("[ERROR] %s: %s" % (cls.get_time(), str(info)))


class JSONParser:
    """ JSON解析器 """
    @classmethod
    def parser(cls, json_path):
        try:
            with open(json_path, 'r', encoding='UTF-8') as f:
                return json.load(f)
        except Exception as e:
            Logger.error(e)
            return {}


def check_env():
    if not os.path.isfile(g_in_excel):
        Logger.error("Excel文件不存在: %s" % g_in_excel)
        exit_delay()
    if not os.path.isfile(g_member_cnf):
        Logger.error("成员列表文件不存在: %s" % g_member_cnf)
        exit_delay()
    if not os.path.isfile(g_version_cnf):
        Logger.error("版本列表文件不存在: %s" % g_version_cnf)
        exit_delay()
    if not os.path.isfile(g_settings_cnf):
        Logger.error("设置项配置文件不存在: %s" % g_settings_cnf)
        exit_delay()
    Logger.info("环境检查完成......成功")


def member_parse():
    global g_member_dict
    g_member_dict = JSONParser.parser(g_member_cnf)
    if not g_member_dict:
        Logger.error("解析成员列表文件失败！")
        exit_delay()
    Logger.info("成员配置解析完成......成功")
    Logger.debug("成员信息: %s" % json.dumps(g_member_dict, ensure_ascii=False, indent=4))


def version_parse():
    global g_include_versions
    global g_exclude_versions
    global g_single_version
    version_dict = JSONParser.parser(g_version_cnf)
    if not version_dict:
        Logger.error("解析版本列表文件失败！")
        exit_delay()
    g_include_versions = version_dict['include']
    g_exclude_versions = version_dict['exclude']
    g_single_version = version_dict['single']
    if not g_single_version:
        Logger.error("single 版本号为空 !")
        exit_delay()
    if not g_include_versions:
        Logger.error("include 版本列表为空 !")
        exit_delay()
    Logger.info("版本配置解析完成......成功")
    Logger.debug("版本信息: %s" % json.dumps(version_dict, indent=4))


def settings_parse():
    global g_root_url
    global g_debug_switch
    global g_change_name
    global g_out_excel
    settings = JSONParser.parser(g_settings_cnf)
    if not settings:
        Logger.error("解析设置项配置文件失败！")
        exit_delay()
    try:
        g_root_url = settings['URL']
        g_debug_switch = settings['DEBUG']
        g_change_name = settings['CH_NAME']
        timestamp = settings['TIMESTAMP']
    except:
        Logger.error("设置项配置文件缺失数据！")
        exit_delay()
    if timestamp:
        g_out_excel = 'DTS-OUT-%s.xlsx' % g_today
    Logger.info("设置项配置解析完成......成功")


class DTSAnalyzer(object):

    def __init__(self):
        # 原始数据帧
        self.all_data = None
        # 未关闭的dts数据帧
        self.dts_data = None
        # 组员外其他人的dts数据帧
        self.other_data = None
        # 组员中英文字典
        self.names_dict = {}
        # 保存各个小组的dts数据帧的字典
        self.group_dict = OrderedDict()
        # DI统计项列表
        self.counter_list = ['归档DI', '开发DI', '总DI']
        record_list = g_di_level + self.counter_list
        # 用于统计汇总界面数据的字典
        self.total_dict = {'类别': record_list}
        # 个人DI排行榜数据结构体的字典
        self.topdi_dict = {'columns': record_list, 'index': [], 'data': []}
        # 每日总DI数据字典
        self.daily_dict = {'Date': []}
        # 每日总DI数据帧
        self.daily_data = None

    @classmethod
    def drop_duplicates(cls, data, key='问题单号', keep='first'):
        data.drop_duplicates(key, keep, inplace=True)   # data 为数据帧地址，直接操作即可

    def read_excel(self):
        self.all_data = pd.read_excel(g_in_excel)
        self.drop_duplicates(self.all_data)

        if os.path.isfile(g_di_excel):
            self.daily_data = pd.read_excel(g_di_excel)       # 第一行数据(日期)作为列名
        else:
            self.daily_data = pd.DataFrame(self.daily_dict)
        Logger.info("读取输入数据完成......成功")

    def version_filter(self):
        # 取出include的版本号
        eval_str = ''
        for version in g_include_versions:
            if eval_str == '':
                eval_str += "self.all_data['B版本'].str.startswith('%s')" % version
            else:
                eval_str += "| self.all_data['B版本'].str.startswith('%s')" % version
        Logger.debug("eval_str for include versions: %s" % eval_str)
        self.all_data = self.all_data[eval(eval_str)]

        # 剔除exclude的版本号
        eval_str = ''
        for version in g_exclude_versions:
            if eval_str == '':
                eval_str += "~ (self.all_data['B版本'].str.startswith('%s'))" % version
            else:
                eval_str += "& (~(self.all_data['B版本'].str.startswith('%s')))" % version
        if eval_str != '':
            Logger.debug("eval_str for exclude versions: %s" % eval_str)
            self.all_data = self.all_data.loc[eval(eval_str)]
        Logger.info("过滤版本号完成......成功")

    def handle_filter(self):
        self.dts_data = self.all_data.loc[self.all_data['当前处理人'].notnull()]
        self.other_data = self.dts_data.loc[self.dts_data['B版本'].str.startswith(g_single_version) == True]
        Logger.info("DTS数据去重完成......成功")

    def di_counter(self, data):
        # python 浮点运算有损，不准确，使用Decimal计算
        filing_di = Decimal(0)
        handle_di = Decimal(0)
        total_di = Decimal(0)
        count_list = []
        # 归档中的DTS
        filing_dts = data[data['当前状态'].isin(g_filing_list)]
        for level in g_di_level:
            count_f = len(filing_dts[filing_dts['严重程度'] == level])
            count_g = len(data[data['严重程度'] == level])
            _lvl_di = Decimal(g_level_map[level])
            filing_di += Decimal(count_f) * _lvl_di
            total_di += Decimal(count_g) * _lvl_di
            count_list.append(count_g)
        handle_di = total_di - filing_di
        return count_list + [filing_di, handle_di, total_di]  # 致命、严重、一般、提示、归档、开发、总共

    # 获取小组DTS数据帧以及测试回归数据帧
    def group_filter(self):
        def data_counter(data, sheet, view=True, add=True):
            self.drop_duplicates(data)
            self.group_dict[sheet] = data
            di_list = self.di_counter(data)
            self.total_dict[sheet] = di_list
            if view:
                if add:
                    self.daily_dict['研发总DI'] += di_list[-1]
                self.daily_dict[sheet] = di_list[-1]

            Logger.info("统计%s DTS......成功" % sheet)
        def is_me(data, key, me):
            return ((data[key] == me) | (data[key].str.contains(me)))

        self.daily_dict['Date'].append(g_today)
        self.daily_dict['研发总DI'] = 0
        regress_data = None
        handled_data = None
        for group, members in g_member_dict.items():
            group_data = None
            for en, ch in members.items():
                self.names_dict[en] = ch
                # 获取该成员处理中的DTS(开发DTS)
                member_dts = self.dts_data.loc[is_me(self.dts_data, '当前处理人', en) & (~self.dts_data['当前状态'].isin(g_regress_list))]
                # 获取测试回归阶段的DTS
                regress_dts = self.dts_data.loc[is_me(self.dts_data, '所有实施修改人', en) & (self.dts_data['当前状态'].isin(g_regress_list))]
                # 获取该成员所有修改的问题单
                handled_dts = self.all_data.loc[is_me(self.all_data, '所有实施修改人', en)]

                group_data = member_dts if group_data is None else pd.concat([group_data, member_dts])
                regress_data = regress_dts if regress_data is None else pd.concat([regress_data, regress_dts])
                handled_data = handled_dts if handled_data is None else pd.concat([handled_data, handled_dts])
                # 统计个人DI数据
                self.topdi_dict['index'].append(ch)
                self.topdi_dict['data'].append(self.di_counter(member_dts))
            # 统计小组DI数据
            data_counter(group_data, group, True, True)
            # 从组员外数据中剔除组员数据和测试回归数据
            self.other_data = self.other_data.loc[(~self.other_data['问题单号'].isin(group_data['问题单号'])) &
                                                  (~self.other_data['问题单号'].isin(regress_data['问题单号']))]
        # 统计测试回归DI数据
        data_counter(regress_data, '测试回归', True, False)
        # 统计组外人员DI数据
        data_counter(self.other_data, '组外其他', True, True)
        # 统计组外人员DI数据
        data_counter(handled_data, '组内修改', False, False)

    def make_plot(self):
        self.daily_data = pd.concat([self.daily_data, pd.DataFrame(self.daily_dict)])
        self.daily_data['Date'] = self.daily_data['Date'].astype('int')
        self.daily_data.iloc[:, 1:] = self.daily_data.iloc[:, 1:].astype('float')
        self.drop_duplicates(self.daily_data, 'Date', 'last')      # 保留最新统计的（最后一个）
        # 数据写入excel
        writer = pd.ExcelWriter(g_di_excel)
        self.daily_data.to_excel(writer, index=False)
        try:
            writer.save()
            writer.close()
        except:
            Logger.info("写入%s失败，请先手动关闭打开的同名文件" % g_di_excel)
            exit_delay(20)
        Logger.info("DI曲线数据写入%s......成功" % g_di_excel)
        # 画图
        plt.rcParams['font.sans-serif'] = ['SimHei']     # 用来正常显示中文标签
        data = pd.read_excel(g_di_excel, parse_dates=True, index_col=0)
        data.plot(title="各组DI曲线图", grid=True, style='o-', figsize=(16, 7))
        plt.savefig('DiData.png', dpi=100)
        #plt.show()

    def write_excel(self):
        def replace_name(data):
            if not g_change_name:
                return
            columns = ["当前处理人", "所有实施修改人"]
            if "创建人" in data.columns:
                columns = columns + ["创建人"]
            for en, ch in self.names_dict.items():
                for clm in columns:
                    data[clm] = data[clm].str.replace(en, ch)
        def create_link(data):
            for _, row in data.iterrows():
                dts_num = row['问题单号']
                row.loc['链接'] = '=HYPERLINK("%s%s", "打开")' % (g_root_url, dts_num)
        def sort_values(in_data):
            return in_data.sort_values(by=['当前处理人', '严重程度', '当前状态'])

        writer = pd.ExcelWriter(g_out_excel)
        for group in self.group_dict.keys():
            group_data = sort_values(self.group_dict[group])
            create_link(group_data)
            replace_name(group_data)
            # 各组sheet页数据写入excel
            group_data.to_excel(writer, sheet_name=group, index=False)
        # DI汇总sheet页数据写入excel
        total_data = pd.DataFrame(self.total_dict)
        total_data.to_excel(writer, sheet_name='DI汇总', index=False)
        # 个人DI排行sheet页数据写入excel
        topdi_data = pd.DataFrame(**self.topdi_dict)
        topdi_data[self.counter_list] = topdi_data[self.counter_list].astype('float')
        topdi_data = topdi_data.sort_values(by=['开发DI'], ascending=[False])
        topdi_data.to_excel(writer, sheet_name='个人DI排行')
        try:
            writer.save()
            writer.close()
        except:
            Logger.info("写入%s失败，请先手动关闭打开的同名文件" % g_out_excel)
            exit_delay(20)
        # 插入可视化图片
        wb = load_workbook(g_out_excel)
        sheet = wb.create_sheet('DI可视化')
        sheet.add_image(Image('DiData.png'), 'A1')
        wb.save(g_out_excel)
        os.remove('DiData.png')
        Logger.info("DTS数据写入%s......成功" % g_out_excel)


def open_excel():
    os.startfile(g_out_excel)
    Logger.info("自动打开%s......成功" % g_out_excel)


def main():
    about_info()
    check_env()
    settings_parse()
    member_parse()
    version_parse()
    analyze = DTSAnalyzer()
    analyze.read_excel()
    analyze.version_filter()
    analyze.handle_filter()
    analyze.group_filter()
    analyze.make_plot()
    analyze.write_excel()
    open_excel()

if __name__ == '__main__':
    try:
        main()
    except:
        Logger.error(traceback.format_exc())
        exit_delay(20)
    exit_delay()

